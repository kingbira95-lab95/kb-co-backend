"""Excel export service using openpyxl."""
import io
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

GOLD_HEX = "D4AF37"
NAVY_HEX = "0A0F1E"
DARK_HEX = "0D1526"
HEADER_FILL = PatternFill("solid", fgColor=NAVY_HEX)
ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
GOLD_FONT = Font(color=GOLD_HEX, bold=True)
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def _header_row(ws, row: int, cols: list[str]) -> None:
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = GOLD_FONT if i == 1 else WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generate_portfolio_excel(
    portfolio_name: str,
    owner_name: str,
    holdings: List[dict],
    prices: dict,
) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Holdings ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Holdings"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:I1")
    ws["A1"] = "KB & Co Corporate Investment Limited"
    ws["A1"].font = Font(color=GOLD_HEX, bold=True, size=14)
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Portfolio: {portfolio_name}  |  {owner_name}  |  {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = Font(color="6B7280", size=9)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16

    headers = ["Symbol", "Shares", "Buy Price (₦)", "Current Price (₦)", "Cost Basis (₦)", "Market Value (₦)", "P&L (₦)", "Return (%)", "Buy Date"]
    _header_row(ws, 4, headers)
    ws.row_dimensions[4].height = 22

    total_cost = total_value = 0.0
    for r, h in enumerate(holdings, start=5):
        sym = h["symbol"]
        shares = h["shares"]
        buy_p = h["buy_price"]
        curr = prices.get(sym, buy_p)
        cost = shares * buy_p
        val = shares * curr
        pnl = val - cost
        ret = (pnl / cost * 100) if cost else 0
        total_cost += cost
        total_value += val

        row_data = [sym, shares, buy_p, curr, cost, val, pnl, ret / 100, h.get("buy_date", "")]
        fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c, val_cell in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val_cell)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left", vertical="center")
            if c in (3, 4, 5, 6):
                cell.number_format = '#,##0.00'
            elif c == 7:
                cell.number_format = '+#,##0.00;-#,##0.00'
                cell.font = Font(color="22C55E" if pnl >= 0 else "EF4444", bold=True)
            elif c == 8:
                cell.number_format = '+0.0%;-0.0%'
                cell.font = Font(color="22C55E" if pnl >= 0 else "EF4444")
        ws.row_dimensions[r].height = 18

    # Total row
    tr = len(holdings) + 5
    total_pnl = total_value - total_cost
    total_ret = (total_pnl / total_cost) if total_cost else 0
    total_row = ["TOTAL", "", "", "", total_cost, total_value, total_pnl, total_ret, ""]
    for c, v in enumerate(total_row, 1):
        cell = ws.cell(row=tr, column=c, value=v)
        cell.font = Font(bold=True, color=GOLD_HEX if c == 1 else "000000")
        cell.fill = PatternFill("solid", fgColor="F0F4F8")
        cell.border = THIN
        cell.alignment = Alignment(horizontal="right" if c > 1 else "left")
        if c in (5, 6):
            cell.number_format = '#,##0.00'
        elif c == 7:
            cell.number_format = '+#,##0.00;-#,##0.00'
            cell.font = Font(bold=True, color="22C55E" if total_pnl >= 0 else "EF4444")
        elif c == 8:
            cell.number_format = '+0.0%;-0.0%'

    _auto_width(ws)
    ws.freeze_panes = "A5"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    summary_data = [
        ("Total Invested", total_cost),
        ("Total Market Value", total_value),
        ("Total P&L", total_pnl),
        ("Overall Return", total_ret),
        ("Number of Holdings", len(holdings)),
        ("Report Generated", datetime.now().strftime("%d %b %Y %H:%M")),
    ]
    ws2["A1"] = "Portfolio Summary"
    ws2["A1"].font = Font(bold=True, size=13, color=NAVY_HEX)
    for i, (label, val) in enumerate(summary_data, start=3):
        ws2.cell(row=i, column=1, value=label).font = BOLD
        cell = ws2.cell(row=i, column=2, value=val)
        if isinstance(val, float) and i <= 5:
            cell.number_format = '#,##0.00' if i != 4 else '+0.00%'
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_stocks_excel(stocks: List[dict]) -> bytes:
    """Export full NGX stocks list to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NGX Stocks"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = f"KB & Co — NGX Equities Price List  |  {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws["A1"].font = Font(color=GOLD_HEX, bold=True, size=12)

    headers = ["Symbol", "Name", "Sector", "Price (₦)", "Change (₦)", "Change (%)", "Volume", "Market Cap (₦M)", "52W High", "52W Low"]
    _header_row(ws, 3, headers)

    for r, s in enumerate(stocks, start=4):
        change = s.get("change", 0)
        chg_pct = s.get("change_pct", 0)
        row = [
            s.get("symbol"), s.get("name"), s.get("sector"),
            s.get("price"), change, chg_pct / 100,
            s.get("volume"), (s.get("market_cap") or 0) / 1e6,
            s.get("high_52w"), s.get("low_52w"),
        ]
        fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(horizontal="right" if c > 3 else "left")
            if c == 4:
                cell.number_format = '#,##0.00'
            elif c == 5:
                cell.number_format = '+#,##0.00;-#,##0.00'
                cell.font = Font(color="22C55E" if change >= 0 else "EF4444")
            elif c == 6:
                cell.number_format = '+0.00%;-0.00%'
                cell.font = Font(color="22C55E" if chg_pct >= 0 else "EF4444")
            elif c in (7, 8, 9, 10):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 18

    _auto_width(ws)
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
